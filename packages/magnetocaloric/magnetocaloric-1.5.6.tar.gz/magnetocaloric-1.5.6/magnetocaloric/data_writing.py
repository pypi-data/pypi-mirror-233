import xlsxwriter
import openpyxl
import os

def data_writing(g_name, file_dir, n, T, Label_one, six_entropy_change_con, M_sqr, one_H_by_M_con, M_pow_MFT, H_by_M_pow_MFT, M_pow_TMFT, H_by_M_pow_TMFT, M_pow_3DH, H_by_M_pow_3DH, M_pow_3DI, H_by_M_pow_3DI, susceptibility_final, T_FWHM_con_final, RCP_final, H_for_RCP):
    """
    Store data to two separate Excel spreadsheets.

    Args:
        n (int): Number of data points.
        T (list): List of temperature values.
        one_H_by_M_con (list): List of lists containing H/M values for each temperature.
        M_sqr (list): List of lists containing M^2 values for each temperature.
        six_entropy_change_con (list): List of lists containing entropy change values for each temperature.

    Note:
        The function will insert appropriate headers and blank spaces in the data before saving to Excel.
    """

    # Specify the folder where you want to save the Excel file
    folder_path = os.path.dirname(file_dir)

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Create a new worksheet
    worksheet = workbook.active

    # Write data to the worksheet
    worksheet['A1'] = 'Hello'
    worksheet['B1'] = 'World'

    new_folder_name = 'mce_output'
    new_folder_path = os.path.join(folder_path, new_folder_name)

    # Check if the folder already exists, and create it if not
    if not os.path.exists(new_folder_path):
        os.makedirs(new_folder_path)
        print(f"</> folder '{new_folder_name}' created at '{new_folder_path}'")
    else:
        print(f"</> folder '{new_folder_name}' already exists at '{new_folder_path}'")

    # Create the full file path including the folder and filename
    file_path01 = os.path.join(folder_path, 'mce_output', 'obtained_delSm_data.xlsx')
    file_path02 = os.path.join(folder_path, 'mce_output', 'obtained_modified_arrott_plot01_data.xlsx')
    file_path03 = os.path.join(folder_path, 'mce_output', 'obtained_modified_arrott_plot02_data.xlsx')
    file_path04 = os.path.join(folder_path, 'mce_output', 'obtained_modified_arrott_plot03_data.xlsx')
    file_path05 = os.path.join(folder_path, 'mce_output', 'obtained_modified_arrott_plot04_data.xlsx')
    file_path06 = os.path.join(folder_path, 'mce_output', 'obtained_susceptibility_data.xlsx')
    file_path07 = os.path.join(folder_path, 'mce_output', 'obtained_RCP.xlsx')

    # Save the workbook to the specified folder and file
    workbook.save(file_path01)
    print ("</> workbook.save ----> [file_path01]")
    workbook.save(file_path02)
    print ("</> workbook.save ----> [file_path02]")
    workbook.save(file_path03)
    print ("</> workbook.save ----> [file_path03]")
    workbook.save(file_path04)
    print ("</> workbook.save ----> [file_path04]")
    workbook.save(file_path05)
    print ("</> workbook.save ----> [file_path05]")
    workbook.save(file_path06)
    print ("</> workbook.save ----> [file_path06]")
    workbook.save(file_path07)
    print ("</> workbook.save ----> [file_path07]")

    # Close the workbook
    workbook.close()

    print ("</> workbook ----> closed")

    # Save entropy change data to the first Excel spreadsheet (path_two).
    workbook = xlsxwriter.Workbook(file_path01)
    worksheet = workbook.add_worksheet()
    row = 0
    six_entropy_change_con[0].insert(0, 'Field(H)')
    six_entropy_change_con[0].insert(1, 'Temperature(T)')
    for i in range (1, len(six_entropy_change_con)):
        six_entropy_change_con[i].insert(0, Label_one[i-1])
        six_entropy_change_con[i].insert(1, '↓')
    if (g_name== 'dSmT_plot' or g_name== 'all_plots'):        
        for col, data in enumerate(six_entropy_change_con):
            worksheet.write_column(row, col, data)
        workbook.close()
    else:
        pass




    # Prepare M^2 vs. H/M data before saving to the second Excel spreadsheet (path_three).
    for i in range(0, 2 * n, 1):
        lo = 2 * i + 1
        T.insert(lo, '   ')


    MFT = H_by_M_pow_MFT.tolist()
    M_pow_MFT_tolist = M_pow_MFT.tolist()

    for i in range(0, n, 1):
        x_index = 2 * i + 1
        MFT.insert(x_index, M_pow_MFT_tolist[i])

    for i in range(0, 2 * n, 1):
        MFT[i].insert(0, T[i])
        MFT[i].insert(1, '↓')
        if i % 2 == 0:
            MFT[i].insert(2, 'H/M')
        else:
            MFT[i].insert(2, 'M^2')

    # Save M^2 vs. H/M data to the second Excel spreadsheet (path_three).
    if (g_name== 'MFT_plot' or g_name== 'all_plots'):
        workbook = xlsxwriter.Workbook(file_path02)
        worksheet = workbook.add_worksheet()
        row = 0
        for col, data in enumerate(MFT):
            worksheet.write_column(row, col, data)
        workbook.close()
    else:
        pass

    if (g_name== 'arrott_plots' or g_name== 'all_plots'):
        
        TMFT = H_by_M_pow_TMFT.tolist()
        M_pow_TMFT_tolist = M_pow_TMFT.tolist()

        for i in range(0, n, 1):
            x_index = 2 * i + 1
            TMFT.insert(x_index, M_pow_TMFT_tolist[i])

        for i in range(0, 2 * n, 1):
            TMFT[i].insert(0, T[i])
            TMFT[i].insert(1, '↓')
            if i % 2 == 0:
                TMFT[i].insert(2, 'H/M')
            else:
                TMFT[i].insert(2, 'M^(1/0.25)')

 
        workbook = xlsxwriter.Workbook(file_path03)
        worksheet = workbook.add_worksheet()
        row = 0
        for col, data in enumerate(TMFT):
            worksheet.write_column(row, col, data)
        workbook.close()    

        D3DH = H_by_M_pow_3DH.tolist()
        M_pow_3DH_tolist = M_pow_3DH.tolist()

        for i in range(0, n, 1):
            x_index = 2 * i + 1
            D3DH.insert(x_index, M_pow_3DH_tolist[i])

        for i in range(0, 2 * n, 1):
            D3DH[i].insert(0, T[i])
            D3DH[i].insert(1, '↓')
            if i % 2 == 0:
                D3DH[i].insert(2, '(H/M)^(1/1.336)')
            else:
                D3DH[i].insert(2, 'M^(1/0.365)')


        workbook = xlsxwriter.Workbook(file_path04)
        worksheet = workbook.add_worksheet()
        row = 0
        for col, data in enumerate(D3DH):
            worksheet.write_column(row, col, data)
        workbook.close()

        D3DI = H_by_M_pow_3DI.tolist()
        M_pow_3DI_tolist = M_pow_3DI.tolist()

        for i in range(0, n, 1):
            x_index = 2 * i + 1
            D3DI.insert(x_index, M_pow_3DI_tolist[i])

        for i in range(0, 2 * n, 1):
            D3DI[i].insert(0, T[i])
            D3DI[i].insert(1, '↓')
            if i % 2 == 0:
                D3DI[i].insert(2, '(H/M)^(1/1.24)')
            else:
                D3DI[i].insert(2, 'M^(1/0.325)')


        workbook = xlsxwriter.Workbook(file_path05)
        worksheet = workbook.add_worksheet()
        row = 0
        for col, data in enumerate(D3DI):
            worksheet.write_column(row, col, data)
        workbook.close()
    else:
        pass
    if (g_name== 'sus_plot' or g_name== 'all_plots'):
        
        workbook = xlsxwriter.Workbook(file_path06)
        worksheet = workbook.add_worksheet()
        
        for row_num, row_data in enumerate(susceptibility_final):
            for col_num, cell_data in enumerate(row_data):
                worksheet.write(row_num, col_num, cell_data)
        workbook.close()
    else:
        pass

    RCP_sheet = [['Magnetic Field','', 'T_FWHM', 'RCP'], ['↓', '', '↓', '↓']]

    
    for i in range(0, len(H_for_RCP)):
        RCP_sheet_prov = []
        RCP_sheet_prov.append(H_for_RCP[i])
        RCP_sheet_prov.append('→')  
        RCP_sheet_prov.append(T_FWHM_con_final[i])
        RCP_sheet_prov.append(RCP_final[i])  
        RCP_sheet.append(RCP_sheet_prov)

    if (g_name== 'RCP_plot' or g_name== 'all_plots'):
        
        workbook = xlsxwriter.Workbook(file_path07)
        worksheet = workbook.add_worksheet()
        
        for row_num, row_data in enumerate(RCP_sheet):
            for col_num, cell_data in enumerate(row_data):
                worksheet.write(row_num, col_num, cell_data)
        workbook.close()
    else:
        pass

    print("</> file accessed \n ---> data obtained ..... written correctly")

    return
