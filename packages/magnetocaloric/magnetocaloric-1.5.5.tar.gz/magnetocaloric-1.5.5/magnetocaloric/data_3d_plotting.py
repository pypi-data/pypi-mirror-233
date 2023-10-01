# Import necessary libraries
import xlrd
import math
import numpy as np
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.pylab import *
from mpl_toolkits.mplot3d import Axes3D
import xlsxwriter
import openpyxl
import os

def mce_3d(path, sheet_index, T_row, H_col, g_name, M_unit, H_unit, T_unit, dpi, save_data):

    if save_data == 'allow':
        folder_path = os.path.dirname(path)

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()

        # Create a new worksheet
        worksheet = workbook.active

        # Write data to the worksheet
        worksheet['A1'] = 'Hello'
        worksheet['B1'] = 'World'

        new_folder_name = 'mce_output'
        new_folder_path = os.path.join(folder_path, new_folder_name)

        # Check if the folder already exists, and create it if not
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)
            print(f"</> folder '{new_folder_name}' created at '{new_folder_path}'")
        else:
            print(f"</> folder '{new_folder_name}' already exists at '{new_folder_path}'")

        # Create the full file path including the folder and filename
        file_path08 = os.path.join(folder_path, 'mce_output', 'obtained_MH_show.xlsx')
        file_path09 = os.path.join(folder_path, 'mce_output', 'obtained_dMdT_show.xlsx')
        file_path10 = os.path.join(folder_path, 'mce_output', 'obtained_dMdH_show.xlsx')
        file_path11= os.path.join(folder_path, 'mce_output', 'obtained_dSm_show.xlsx')


        # Save the workbook to the specified folder and file
        workbook.save(file_path08)
        print ("</> workbook.save ----> [file_path08]")
        workbook.save(file_path09)
        print ("</> workbook.save ----> [file_path09]")
        workbook.save(file_path10)
        print ("</> workbook.save ----> [file_path10]")
        workbook.save(file_path11)
        print ("</> workbook.save ----> [file_path11]")


        # Close the workbook
        workbook.close()

        print ("</> workbook ----> closed")






    def column_letter_to_number(column_letter):
        column_number = 0
        for char in column_letter:
            # Convert the character to its ASCII value and subtract 64
            # to get the 1-based index (A=1, B=2, ..., Z=26)
            column_number = column_number * 26 + (ord(char.upper()) - 64)
        return column_number

    # Open the Excel workbook
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(sheet_index-1)

    # Get the number of rows and columns in the Excel sheet
    n_rows = sheet.nrows
    n_cols = sheet.ncols

    # Initialize data containers
    init_data = []

    # Initialize lists for temperature and applied field values
    prov_T = []
    for i in range(n_cols):
        prov_T.append(sheet.cell_value((int(T_row)-1), i))

    # Populate init_data with the Excel sheet data
    for i in range(n_cols):
        col_data = []
        for j in range(1, n_rows):
            col_data.append(sheet.cell_value(j, i))
        init_data.append(col_data)

    # Define a function to check if the last element of init_data is a string
    def check_last_element(init_data):
        last_element = init_data[0][-1]
        if isinstance(last_element, str):
            return True
        else:
            return False

    H_col_num = column_letter_to_number(H_col)
    # Remove the first element from prov_T, init_data to store field in H
    prov_T.pop(0)
    H = init_data.pop(int(H_col_num) -1)
    T = prov_T
    M = init_data

    # If the last element of init_data is a string, remove the last element from each list
    if check_last_element(init_data):
        for i in range(0, len(init_data)):
            init_data[i].pop(len(init_data[i]) - 1)
        H.pop(len(H) - 1)
    else:
        pass

    px = int((dpi)**(1/2))
    px_1 = int(px + 1) 
    # Create linearly spaced arrays for applied field (H_lin) and temperature (T_lin)
    H_lin = np.linspace(H[0], H[len(H) - 1], px_1)
    T_lin = np.linspace(T[0], T[len(T) - 1], px_1)

    # Initialize containers for interpolated data
    M_res_H = []
    M_res_HT = []

    # Interpolate data for applied field
    for i in range(0, len(M)):
        pri_M_res_H = np.interp(H_lin, H, M[i])
        nor_pri_M_res_H = pri_M_res_H.tolist()
        M_res_H.append(nor_pri_M_res_H)

    # Interpolate data for applied field and temperature
    for i in range(0, len(H_lin)):
        arrg_M_res_T = []
        for j in range(0, len(M)):
            arrg_M_res_T.append(M_res_H[j][i])
        pri_M_res_HT = np.interp(T_lin, T, arrg_M_res_T)
        nor_pri_M_res_HT = pri_M_res_HT.tolist()
        M_res_HT.append(nor_pri_M_res_HT)


    del_H = H_lin[0] - H_lin[1]
    entropy_change_con = []
    pri03_entropy_change_con = np.zeros(px)
    dMdT_con = []
    pri03_dMdT_con = np.zeros(px)

    # Calculate entropy change and derivative data
    for i in range(0, len(H_lin) - 1):
        pri02_entropy_change_con = []
        pri02_dMdT_con = []
        for j in range(0, len(T_lin) - 1):
            pri01_entropy_change = abs((float(M_res_HT[i][j + 1]) - float(M_res_HT[i][j])) / (float(T_lin[j + 1]) - float(T_lin[j]))) * del_H * 10 ** (-4)
            pri01_dMdT = abs((float(M_res_HT[i][j + 1]) - float(M_res_HT[i][j])) / (float(T_lin[j + 1]) - float(T_lin[j]))) * 10 ** (-4)

            pri02_entropy_change_con.append(pri01_entropy_change)
            pri02_dMdT_con.append(pri01_dMdT)

        pri03_entropy_change_con = pri03_entropy_change_con + pri02_entropy_change_con
        pri03_dMdT_con = pri02_dMdT_con
        nor_pri03_entropy_change_con = pri03_entropy_change_con.tolist()
        nor_pri03_dMdT_con = pri03_dMdT_con
        entropy_change_con.append(nor_pri03_entropy_change_con)
        dMdT_con.append(nor_pri03_dMdT_con)

    # Correct the order of entropy and derivative data
    entropy_change_con_correc = []
    dMdT_con_correc = []
    for i in range(0, len(M_res_HT)):
        corr_ind = len(entropy_change_con) - 1 - i
        entropy_change_con_correc.append(entropy_change_con[corr_ind])
        dMdT_con_correc.append(dMdT_con[corr_ind])

    # Initialize container for derivative data with respect to applied field
    dMdH_con = []
    pri03_dMdH_con = np.zeros(px)

    # Calculate derivative data with respect to applied field
    for j in range(0, len(T_lin) - 1):
        pri02_dMdH_con = []
        for i in range(0, len(H_lin) - 1):
            pri01_dMdH = abs(((float(M_res_HT[i + 1][j])) - (float(M_res_HT[i][j])))/del_H)

            pri02_dMdH_con.append(pri01_dMdH)

        pri03_dMdH_con = pri02_dMdH_con
        nor_pri03_dMdH_con = pri03_dMdH_con
        dMdH_con.append(nor_pri03_dMdH_con)

    # Transpose and convert derivative data with respect to applied field
    dMdH_con_correc_01 = np.transpose(dMdH_con)
    dMdH_con_correc_02 = dMdH_con_correc_01.tolist()
    dMdH_con_correc = []

    # Correct the order of derivative data with respect to applied field
    for i in range(0, len(M_res_HT)):
        corr_ind = len(entropy_change_con) - 1 - i
        dMdH_con_correc.append(dMdH_con_correc_02[corr_ind])


    if (g_name == 'MH_show'):        
        # Data for the first plot
        T_lin, H_lin = np.meshgrid(T_lin, H_lin)
        np_M_res_HT_correc = np.asarray(M_res_HT)

        # Create the first plot
        fig = plt.figure()
        ax1 = fig.add_subplot(111, projection='3d')

        # Plot the surface for the first dataset
        surface1 = ax1.plot_surface(T_lin, H_lin, np_M_res_HT_correc, cmap='plasma')
        ax1.set_xlabel(f'T {T_unit}', fontname="monospace")
        ax1.set_ylabel(f'H {H_unit}', fontname="monospace")
        ax1.set_zlabel(f'M {M_unit}', fontname="monospace")
        ax1.set_title('M distribution', fontname="monospace")
        cb = fig.colorbar(surface1, ax=ax1, label='M Value')
        ax = cb.ax
        text = ax.yaxis.label
        font = matplotlib.font_manager.FontProperties(family='monospace')
        text.set_font_properties(font)
        print ("</> request for MH_show ----> accepted & generated using the corrected form ([np_M_res_HT_correc]) with cmap = 'plasma'")
        # Show the first plot
        plt.show()

        if save_data == 'allow':
            T_for_sheet = T_lin[0].tolist()
            H_for_sheet = []
            for i in range(0, len(H_lin)):
                H_for_sheet.append(H_lin[i][0])
            load_data = []
            T_for_sheet.insert(0, 'Temperature(T)')
            T_for_sheet.insert(1, '')
            ss_lst = ['↓'] * len(T_lin[0])
            ss_lst.insert(0, 'Field(H)')
            ss_lst.insert(1, '')

            load_data.append(T_for_sheet)
            load_data.append(ss_lst)

            
            for i in range(0, len(np_M_res_HT_correc)):
                load_data_prov = []
                load_data_prov = np_M_res_HT_correc[i].tolist()
                load_data_prov.insert(0, H_for_sheet[i])
                load_data_prov.insert(1, '→')  

                load_data.append(load_data_prov)              

            workbook = xlsxwriter.Workbook(file_path08)
            worksheet = workbook.add_worksheet()
            
            for row_num, row_data in enumerate(load_data):
                for col_num, cell_data in enumerate(row_data):
                    worksheet.write(row_num, col_num, cell_data)
            workbook.close()
            print("</> file accessed \n ---> data obtained ..... written correctly")
        else:
            pass


    elif(g_name == 'dSm_show'):    
        # Data for the second plot
        T_lin = np.delete(T_lin, -1)
        H_lin = np.delete(H_lin, -1)
        T_lin, H_lin = np.meshgrid(T_lin, H_lin)
        np_entropy_change_con_correc = np.asarray(entropy_change_con)

        # Create the second plot
        fig = plt.figure()
        ax2 = fig.add_subplot(111, projection='3d')

        # Plot the surface for the second dataset
        surface2 = ax2.plot_surface(T_lin, H_lin, np_entropy_change_con_correc, cmap='twilight')
        ax2.set_xlabel(f'T {T_unit}', fontname="monospace")
        ax2.set_ylabel(f'H {H_unit}', fontname="monospace")
        ax2.set_zlabel(f'∆Sm ({M_unit}).{H_unit}/{T_unit}', fontname="monospace")
        ax2.set_title('∆Sm distribution', fontname="monospace")
        cb = fig.colorbar(surface2, ax=ax2, label='∆Sm Value')
        ax = cb.ax
        text = ax.yaxis.label
        font = matplotlib.font_manager.FontProperties(family='monospace')
        text.set_font_properties(font)
        print ("</> request for dSm_show ----> accepted & generated using the corrected form ([np_entropy_change_con_correc]) with cmap = 'twilight'")
        # Show the second plot

        plt.show()

        if save_data == 'allow':
            T_for_sheet = T_lin[0].tolist()
            H_for_sheet = []
            for i in range(0, len(H_lin)):
                H_for_sheet.append(H_lin[i][0])
            load_data = []
            T_for_sheet.insert(0, 'Temperature(T)')
            T_for_sheet.insert(1, '')
            ss_lst = ['↓'] * len(T_lin[0])
            ss_lst.insert(0, 'Field(H)')
            ss_lst.insert(1, '')

            load_data.append(T_for_sheet)
            load_data.append(ss_lst)

            
            for i in range(0, len(np_entropy_change_con_correc)):
                load_data_prov = []
                load_data_prov = np_entropy_change_con_correc[i].tolist()
                load_data_prov.insert(0, H_for_sheet[i])
                load_data_prov.insert(1, '→')  

                load_data.append(load_data_prov)              

            workbook = xlsxwriter.Workbook(file_path11)
            worksheet = workbook.add_worksheet()
            
            for row_num, row_data in enumerate(load_data):
                for col_num, cell_data in enumerate(row_data):
                    worksheet.write(row_num, col_num, cell_data)
            workbook.close()
            print("</> file accessed \n ---> data obtained ..... written correctly")
        else:
            pass


    elif(g_name == 'dMdH_show'):    
        # Data for the third plot
        T_lin = np.delete(T_lin, -1)
        H_lin = np.delete(H_lin, -1)
        T_lin, H_lin = np.meshgrid(T_lin, H_lin)
        np_dMdH_con_correc = np.asarray(dMdH_con_correc_02)

        # Create the third plot
        fig = plt.figure()
        ax3 = fig.add_subplot(111, projection='3d')

        # Plot the surface for the second dataset
        surface3 = ax3.plot_surface(T_lin, H_lin, np_dMdH_con_correc, cmap='RdBu')
        ax3.set_xlabel(f'T {T_unit}', fontname="monospace")
        ax3.set_ylabel(f'H {H_unit}', fontname="monospace")
        ax3.set_zlabel(f'dM/dH ({M_unit})/{H_unit}', fontname="monospace")
        ax3.set_title('dM/dH distribution', fontname="monospace")
        cb = fig.colorbar(surface3, ax=ax3, label='dM/dH Value')
        ax = cb.ax
        text = ax.yaxis.label
        font = matplotlib.font_manager.FontProperties(family='monospace')
        text.set_font_properties(font)
        print ("</> request for dMdH_show ----> accepted & generated using the corrected form ([np_dMdH_con_correc]) with cmap = 'RdBu'")
        # Show the third plot
        plt.show()

        if save_data == 'allow':
            T_for_sheet = T_lin[0].tolist()
            H_for_sheet = []
            for i in range(0, len(H_lin)):
                H_for_sheet.append(H_lin[i][0])
            load_data = []
            T_for_sheet.insert(0, 'Temperature(T)')
            T_for_sheet.insert(1, '')
            ss_lst = ['↓'] * len(T_lin[0])
            ss_lst.insert(0, 'Field(H)')
            ss_lst.insert(1, '')

            load_data.append(T_for_sheet)
            load_data.append(ss_lst)

            
            for i in range(0, len(np_dMdH_con_correc)):
                load_data_prov = []
                load_data_prov = np_dMdH_con_correc[i].tolist()
                load_data_prov.insert(0, H_for_sheet[i])
                load_data_prov.insert(1, '→')  

                load_data.append(load_data_prov)              

            workbook = xlsxwriter.Workbook(file_path10)
            worksheet = workbook.add_worksheet()
            
            for row_num, row_data in enumerate(load_data):
                for col_num, cell_data in enumerate(row_data):
                    worksheet.write(row_num, col_num, cell_data)
            workbook.close()
            print("</> file accessed \n ---> data obtained ..... written correctly")
        else:
            pass

    elif(g_name == 'dMdT_show'):    
        # Data for the fourth plot
        T_lin = np.delete(T_lin, -1)
        H_lin = np.delete(H_lin, -1)
        T_lin, H_lin = np.meshgrid(T_lin, H_lin)
        np_dMdT_con_correc = np.asarray(dMdT_con)

        # Create the fourth plot
        fig = plt.figure()
        ax4 = fig.add_subplot(111, projection='3d')

        # Plot the surface for the second dataset
        surface4 = ax4.plot_surface(T_lin, H_lin, np_dMdT_con_correc, cmap='BrBG')
        ax4.set_xlabel(f'T {T_unit}', fontname="monospace")
        ax4.set_ylabel(f'H {H_unit}', fontname="monospace")
        ax4.set_zlabel(f'dM/dT ({M_unit})/{T_unit}', fontname="monospace")
        ax4.set_title('dM/dT distribution', fontname="monospace")
        cb = fig.colorbar(surface4, ax=ax4, label='dM/dT Value')
        ax = cb.ax
        text = ax.yaxis.label
        font = matplotlib.font_manager.FontProperties(family='monospace')
        text.set_font_properties(font)
        print ("</> request for dMdT_show ----> accepted & generated using the corrected form ([np_dMdT_con_correc]) with cmap = 'BrBG'")


        #print ("T_lin", T_lin,"\nH_lin", H_lin,"\ndM/dT", np_dMdT_con_correc)
        # Show the fourth plot
        plt.show()

        if save_data == 'allow':
            T_for_sheet = T_lin[0].tolist()
            H_for_sheet = []
            for i in range(0, len(H_lin)):
                H_for_sheet.append(H_lin[i][0])
            load_data = []
            T_for_sheet.insert(0, 'Temperature(T)')
            T_for_sheet.insert(1, '')
            ss_lst = ['↓'] * len(T_lin[0])
            ss_lst.insert(0, 'Field(H)')
            ss_lst.insert(1, '')

            load_data.append(T_for_sheet)
            load_data.append(ss_lst)

            
            for i in range(0, len(np_dMdT_con_correc)):
                load_data_prov = []
                load_data_prov = np_dMdT_con_correc[i].tolist()
                load_data_prov.insert(0, H_for_sheet[i])
                load_data_prov.insert(1, '→')  

                load_data.append(load_data_prov)              

            workbook = xlsxwriter.Workbook(file_path09)
            worksheet = workbook.add_worksheet()
            
            for row_num, row_data in enumerate(load_data):
                for col_num, cell_data in enumerate(row_data):
                    worksheet.write(row_num, col_num, cell_data)
            workbook.close()
            print("</> file accessed \n ---> data obtained ..... written correctly")
        else:
            pass

    else:
        print ("xxx ---> g_name not found")
    
    return ''