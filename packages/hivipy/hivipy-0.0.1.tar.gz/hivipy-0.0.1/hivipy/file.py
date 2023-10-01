from copy import deepcopy
import logging
from operator import inv
import traceback
import os
import sys
from random import *
import json
from datetime import datetime
from dateutil.parser import parse
import pytz
import csv
from urllib.parse import urlparse
from urllib.parse import parse_qs
from pyexcel_xls import get_data as xls_get, save_data as xls_save_data
from pyexcel_xlsx import get_data as xlsx_get, save_data as xlsx_save_data
import codecs
from functools import reduce
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
import base64
import mimetypes
from flask_weasyprint import HTML, CSS, render_pdf
import tempfile
from datauri import DataURI
import qrcode
import qrcode.image.svg
# from barcode import EAN13
# from barcode.writer import ImageWriter
from io import StringIO, BytesIO
from base64 import b64encode
import re
import pandas as pd
from flask import make_response, jsonify, render_template, send_file
from flask.wrappers import Response
from pathlib import Path

from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from .objects import pdDataframeToObject

from .config import langs, dateFormatForFile, eseDatas, responsesPossibles
from .utils import getLang, convertToOtherType, CleanName, GetStaticFile
from .string import RandomStr
from .lang_datas import langDatas as langdatas
from .hivi_init import Manager, createDir, createStaticMediaPath
from .hivi_init import Manager


manager = Manager()
structConf = manager.getStructConfig()
DEBUG = structConf['debug']
log = logging.getLogger(__name__)
export_types = ['csv', 'excel', 'pdf', 'json']

def getUrlDatas(urlStr: str):
    noDatas = {
        'domain': None,
        'port': None,
        'path': None,
        'hostname': None,
        'port': None,
        'query': {},
    }
    try:
        if(
            type(urlStr) == str and
            len(urlStr) > 0
        ):
            datas = urlparse(urlStr)
            return {
                'domain': datas.netloc,
                'path': datas.path,
                'hostname': datas.hostname,
                'port': datas.port,
                'query': (parse_qs(datas.query, keep_blank_values=True)),
            }
        return noDatas
        
    except Exception as err:
        # if DEBUG :
            # print(err)
        stack = str(traceback.format_exc())
        log.error(stack)
        return noDatas

def cleanFilenameWithoutExtension(filename: str, mapCF = lambda x: x, premap = lambda x: x):
    if(filename is not None):
        sep = '_'
        res = (
            premap(
                sep.join(
                    list(
                        filter(
                            lambda x: len(x) > 0,
                            re.sub(
                                re.compile(r"[^a-zA-Z0-9_]", re.MULTILINE),
                                sep,
                                filename,
                            ).split(sep),
                        )
                    )
                )
            ) if callable(premap) else (
                sep.join(
                    list(
                        filter(
                            lambda x: len(x) > 0,
                            re.sub(
                                re.compile(r"[^a-zA-Z0-9_]", re.MULTILINE),
                                sep,
                                filename,
                            ).split(sep),
                        )
                    )
                )
            )
        ) if len(filename) > 0 else None
        res = (
            mapCF(
                res
            ) if callable(mapCF) else res
        ) if len(filename) > 0 else None
        return res
    else:
        return None

def cleanFilename(filename: str, mapCF = lambda x: x, premap = lambda x: x):
    filename = filename if type(filename) == str else None
    resCF = filename
    if(filename is not None):
        sep = '_'
        filename_array = filename.split(".")
        extension = filename_array[-1] if len(filename_array) > 1 else None
        filename_sub = cleanFilenameWithoutExtension(filename=filename_array[:-1][0], mapCF=mapCF, premap = premap)

        resCF = filename_sub + '.' + extension

        # if DEBUG :
            # print("> scripts - file | cleanFilename - extension:: ", extension)
            # print("> scripts - file | cleanFilename - filename_sub:: ", filename_sub)
            # print("> scripts - file | cleanFilename - resCF:: ", resCF)

    return resCF


def createPath(*paths):
    """ createPath - Fonction de creation d'un chemin absolu

    Cette methode permet de creer un chemin absolu en fonction de plusieurs autres chemins

    Parameters
    ------------
        paths: list
            l'ensemble de tous les paths que vous ciblez pour votre chemin
    Return
    -----------
        dir : str
            Le path cree en fonction des autres paths
    """
    if(type(paths) in (list, tuple) and paths is not None):
        paths = list(
            map(
                lambda path: str(path),
                list(
                    filter(
                        lambda path: type(path) in (int, float) or ( type(path) == str and len(path) > 0 ),
                        paths,
                    )
                )
            )
        )
        for index, path in enumerate(paths):
            if(index < len(paths) - 1):
                paths[index] = paths[index].replace('.', '\\')
        res = str(os.path.join(*paths).replace('/', '\\').replace('\\\\', '\\'))
        res = res if res != '\\' else ''
        return res
    else:
        return None

def createFolderTree(
    folderPath: str
):
    folderPath = folderPath if type(folderPath) in (str, int, float) and len(str(folderPath)) > 0 else None
    if(folderPath is not None):
        folderPath = createPath(folderPath)

        allPathDest = list(
            filter(
                lambda path: type(path) == str and len(path) > 0,
                folderPath.split('\\'),
            )
        )
        for i in range(1, len(allPathDest) + 1):
            path = createStaticMediaPath('\\'.join(allPathDest[:i]))
            if(len(path) > 0 and not(os.path.exists(path) == True)):
                os.mkdir(path)

    return folderPath


def ConvertForBase64File(file):
    result = None
    if(
        type(file) in (list, tuple)
        and len(file) > 0
    ) :
        result = str(''.join(file))
    elif(
        type(file) == str
        and len(file) > 0
    ) :
        result = str(file)
    return result
def RetroConvertForBase64File(file):
    # len_max = 256
    len_max = 100000
    result = []
    if(
        type(file) == str
        and len(file) > 0
    ) :
        result = tuple(
            map(
                lambda x: file[x['e1']:x['e2']],
                tuple(
                    map(
                        lambda x: {'e1': x, 'e2': x+len_max},
                        tuple(range(0, len(file), len_max))
                    )
                )
            )
        )
    elif (
        type(file) in (list, tuple)
    ):
        result = file
    return result


def checkIfBASE64(value):
    if type(value) == str and len(value) > 0:
        paddingRequired = True
        urlSafe = True
        nbr1 = '=' if paddingRequired else '(={0,1})'
        nbr2 = '(\-{1})' if urlSafe else '(+{1})'
        nbr3 = '(_{1})' if urlSafe else '(\\{1})'

        ruleValue: str = r"^(?:[A-Za-z0-9" + nbr2 + "/]{4})*(?:[A-Za-z0-9" + nbr2 + "/]{2}==|[A-Za-z0-9" + nbr2 + "/]{3}" + nbr1 + "|[A-Za-z0-9" + nbr2 + "/]{4})$"
        flag: re.RegexFlag = (re.MULTILINE)
        flag = flag if (
            type(flag) is re.RegexFlag and
            flag is not None
        ) else None
        ruleValue = ruleValue if (
            type(ruleValue) == str and
            len(ruleValue) > 0
        ) else ''
        if flag is not None:
            ruleValue = re.compile(ruleValue, flag)

        return bool(re.match(ruleValue, str(value)))
    return False
def dataURIGetFIleData(
    data: str
):
    data = deepcopy(ConvertForBase64File(data))
    dest = 'tmp'
    try:
        uri = DataURI(data)
        if(checkIfBASE64(data)):
            result = None
        elif(uri.data):
            mimeType = uri.mimetype
            extension = mimetypes.guess_extension(mimeType) if mimeType is not None else None
            header_encoded_data, encoded_data = data.split(",", 1)
            
            dest = createFolderTree(dest)
            finalUrlDest = createStaticMediaPath(dest)
            
            dest = createFolderTree(dest)
            newFileName = 'tmp' + RandomStr(
                lengthStr=30
            )
            finalUrlDest = createStaticMediaPath(dest)
            finalUrlFile = createStaticMediaPath(dest, newFileName)
            finalUrlFilePartial = createPath(dest, newFileName)

            fileData = None
            file_content=base64.b64decode(encoded_data)
            with open(finalUrlFile,"wb") as f:
                f.write(
                    file_content
                )
            statFile = Path(finalUrlFile).stat()

            # if DEBUG :
                # print('--> hivipy - fileData:: ', fileData)
            # fileNameRes = Upload(fileData, 'base64saver')

            result = {
                'mimetype': mimeType,
                'extension': extension,
                'size': statFile.st_size
            }

            if os.path.exists(finalUrlFile):
                os.remove(finalUrlFile)

            # dest = createFolderTree(dest)
            # result = finalUrlFilePartial
        else:
            result = None
    except:        
        logging.getLogger("error_logger").error(traceback.format_exc())
        result = None
    return result
def dataURIUpload(
    data: str,
    dest: str = 'tmp',
):
    data = deepcopy(ConvertForBase64File(data))
    dest = deepcopy(dest) if type(dest) == str and len(dest) > 0 else 'tmp'
    try:
        uri = DataURI(data)
        if(uri.data):
            mimeType = uri.mimetype
            extension = mimetypes.guess_extension(mimeType) if mimeType is not None else None
            header_encoded_data, encoded_data = data.split(",", 1)
            
            dest = createFolderTree(dest)
            newFileName = (
                RandomStr(
                    lengthStr=30
                ) + extension
            )
            finalUrlDest = createStaticMediaPath(dest)
            finalUrlFile = createStaticMediaPath(dest, newFileName)
            finalUrlFilePartial = createPath(dest, newFileName)

            fileData = None
            file_content=base64.b64decode(encoded_data)
            # print("> dataURIUpload - file_content::", file_content)
            # print("> dataURIUpload - dest::", dest)
            # print("> dataURIUpload - newFileName::", newFileName)
            # print("> dataURIUpload - finalUrlDest::", finalUrlDest)
            # print("> dataURIUpload - finalUrlFile::", finalUrlFile)
            # print("> dataURIUpload - finalUrlFilePartial::", finalUrlFilePartial)

            with open(finalUrlFile,"wb") as f:
                f.write(
                    file_content
                )

            # if DEBUG :
                # print('--> hivipy - fileData:: ', fileData)
            # fileNameRes = Upload(fileData, 'base64saver')

            result = finalUrlFilePartial

            # fp = tempfile.TemporaryFile()
            # fp.write(b'Hello world!')
            # fp.seek(0)
            # fp.read()
            # fp.close()

            # dest = createFolderTree(dest)
            # result = finalUrlFilePartial
        else:
            result = None
    except:        
        logging.getLogger("error_logger").error(traceback.format_exc())
        result = None
    return result
def Upload(
    file: FileStorage,
    dest: str = '',
    winMessage: dict = {
        'type': 'success',
        'message' : {
            'fr': 'file uploadé avec succès',
            'en': 'file uploaded successfully',
        },
    },
    loseMessage: dict = {
        'type': 'danger',
        'message' : {
            'fr': 'echec lors de l\'upload du file du file',
            'en': 'failure while uploading file from file',
        },
    },
    showMessage = True,
    lang = 'fr',
    map = lambda x: "{filename}{sub}{sub2}".format(
        filename = x,
        sub = RandomStr(
            lengthStr=10
        ),
        sub2 = datetime.now().strftime(dateFormatForFile)
    ),
):
    def cleanMessage(message: str):
        resCM = {
            'type': message['type'],
            'message': message['message'][lang],
        }
        return resCM
    lang = getLang(lang)
    showMessage = showMessage if type(showMessage) == bool else True
    result = loseMessage if showMessage else False
    try:
        map = map if callable(map) else (
            lambda x: "{filename}{sub}{sub2}".format(
                filename = x,
                sub = RandomStr(
                    lengthStr=10
                ),
                sub2 = datetime.now().strftime(dateFormatForFile)
            )
        )
        dest = createFolderTree(dest)
        print("> Upload - file::", file)
        if DEBUG :
            # print("> Upload - file - map::", map)
            print("> Upload - file::", file)

        newFileName = cleanFilename(secure_filename(file.filename), mapCF = map, premap = lambda x: x[:10])
        finalUrlFilePartial = createPath(dest, newFileName)
        finalUrlFile = createStaticMediaPath(dest, newFileName)
        # if DEBUG :
            # print("> Upload - file::", file)
            # print("> Upload - file - contentType::", contentType)
            # print("> Upload - file - fileExtension::", fileExtension)
            # print("> Upload - file - oriFileName::", oriFileName)
            # print("> Upload - file - newFileName::", newFileName)
            # print("> Upload - file - file.filename::", file.filename)
            # print("> Upload - file - newFileName::", newFileName)
            # print("> Upload - file - dest::", dest)
            # print("> Upload - file - finalUrlFilePartial::", finalUrlFilePartial)
            # print("> Upload - file - finalUrlFile::", finalUrlFile)
            # print("> Upload - file - min_sizeAuth::", min_sizeAuth)
            # print("> Upload - file - max_sizeAuth::", max_sizeAuth)

        file.seek(0)
        file.save(finalUrlFile)
        result = finalUrlFilePartial
    except:        
        logging.getLogger("error_logger").error(traceback.format_exc())
        result = None
    return result
def Download(
    fileName,
    newFileName = None,
    src = '',
    loseMessage = {
        'type': responsesPossibles['unknown_error']['type'],
        'code': responsesPossibles['unknown_error']['code'],
        'status': responsesPossibles['unknown_error']['status'],
        'message' : 'Echec lors du telechargement du fichier',
    },
    returnException: bool = True,
    map = lambda x: "{sub2}{filename}{sub}".format(
        filename = x,
        sub = RandomStr(
            lengthStr=10
        ),
        sub2 = datetime.now().strftime(dateFormatForFile)
    ),
) -> Response:
    map = map if callable(map) else (
        lambda x: "{sub2}{filename}{sub}".format(
            filename = x,
            sub = RandomStr(
                lengthStr=10
            ),
            sub2 = datetime.now().strftime(dateFormatForFile)
        )
    )
    returnException = deepcopy(returnException) if type(returnException) == bool else False
    loseMessage = loseMessage if ( type(loseMessage) == dict and 'type' in list(loseMessage.keys()) and type(loseMessage['type']) == str and len(loseMessage['type']) > 0 and 'message' in list(loseMessage.keys()) and type(loseMessage['message']) == str and len(loseMessage['message']) > 0 ) else {
        'type': responsesPossibles['unknown_error']['type'],
        'code': responsesPossibles['unknown_error']['code'],
        'status': responsesPossibles['unknown_error']['status'],
        'message' : 'Echec lors du telechargement du fichier',
    }
    result = {}
    try:
        newFileName = newFileName if type(newFileName) == str and len(newFileName) > 0 else fileName
        finalUrlFile = createStaticMediaPath(src, newFileName)
        newFileName = cleanFilename(secure_filename(fileName), mapCF = map, premap = lambda x: x[:50])
        # if DEBUG :
        #     print("> file - Download | finalUrlFile:: ", finalUrlFile)
        #     print("> file - Download | newFileName:: ", newFileName)
        return send_file(
            finalUrlFile,
            as_attachment=True,
            attachment_filename=newFileName,
            download_name=newFileName,
        )
    except Exception as err:
        code = str(type(err))
        msg = str(err)
        stack = str(traceback.format_exc())
        trace = sys.exc_info()[2]
        
        if DEBUG == True:
            log.error(stack)

        logging.getLogger("error_logger").error(traceback.format_exc())

        stackError = stack if DEBUG else None
        loseMessage['stack'] = stackError

        result = jsonify(loseMessage)
    return result
def EncodeFile(image_path, encoder = 'utf-8'):
    result = None
    encoders = ['utf-8', 'ascii']
    encoder = encoder if encoder in encoders else encoders[0]
    try:
        with open(image_path, "rb") as image_file:
            contentType = mimetypes.guess_type(image_path)
            contentType = contentType[0] if contentType[0] else contentType
            result = base64.b64encode(image_file.read()).decode(encoder)
            result = 'data:{contentType};base64,{image_data}'.format(
                contentType = contentType,
                image_data = result
            )
    except:        
        logging.getLogger("error_logger").error(traceback.format_exc())
        result = None
    return result

def Export_genQRCode(
    data,
    method = 'basic',
):
    svg_out = BytesIO()
    method = method if method in ['basic', 'fragment', 'path'] else 'basic'
    if method == 'basic':
        factory = qrcode.image.svg.SvgImage
    elif method == 'fragment':
        factory = qrcode.image.svg.SvgFragmentImage
    elif method == 'path':
        factory = qrcode.image.svg.SvgPathImage
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=15,
        image_factory=factory
    )
    qr.add_data(data)
    img = qr.make_image()
    img.save(svg_out)
    result = base64.b64encode(svg_out.getvalue()).decode()
    result = "data:image/svg+xml;charset=utf-8;base64," + result
    return result
def Export_genBarCode(
    data,
):
    svg_out = BytesIO()
    code = EAN13(data, writer=ImageWriter())
    code.write(svg_out)

    result = svg_out.getvalue()
    #result = base64.b64encode(result).decode()
    #result = "data:image/svg+xml;charset=utf-8;base64," + result
    return result
def Export__initFilename(filename, export_type = 'csv'):
    export_type = export_type if export_type in export_types else export_types[0]
    filetypes = {
        'excel': 'xlsx',
        'csv': 'csv',
        'pdf': 'pdf',
        'json': 'json',
    }
    filetype = filetypes[export_type] if export_type in list(filetypes.keys()) else filetypes[export_types[0]]
    dateAct = datetime.now(tz=pytz.UTC)
    filename = cleanFilenameWithoutExtension(
        filename=filename,
        mapCF=lambda x: "{data}_{sup1}_{sup2}".format(
            data = x,
            sup1 = dateAct.strftime('%Y%m%d%H%M%S'),
            sup2 = RandomStr(lengthStr = 20),
        )
    )
    filename = "{0}.{1}".format(
        filename,
        filetype
    )
    return filename
def Export__initColumnsConf(columnsConf):
    result = []
    columnsConf = columnsConf if type(columnsConf) == list or type(columnsConf) == tuple else []
    for columnConf in columnsConf:
        columnConf = columnConf if type(columnConf) == dict else {}
        name = columnConf['name'] if (
            'name' in columnConf.keys() and
            type(columnConf['name']) == str and
            len(columnConf['name']) > 0
        ) else None
        label = columnConf['label'] if (
            'label' in columnConf.keys() and
            type(columnConf['label']) == str and
            len(columnConf['label']) > 0
        ) else None
        pos = columnConf['pos'] if (
            'pos' in columnConf.keys() and
            type(columnConf['pos']) == int
        ) else None
        if name != None and label != None and pos != None:
            result.append({
                'name': name,
                'label': label,
                'pos': pos,
            })
    result = sorted(
        result,
        key=lambda k: k['pos']
    )
    return result
def Export__initColumns(columnsConf):
    columns = list(
        map(
            lambda x: x['label'],
            columnsConf,
        )
    )
    columns = list(
        map(
            lambda x: x.capitalize(),
            columns,
        )
    )
    return columns
def Export__initRows(rows, columnsConf):
    # if DEBUG :
        # print("> file - Export__initRows | rows:: ", rows)
        # print("> file - Export__initRows | columnsConf:: ", columnsConf)
    def getDatas(data):
        data = data if type(data) == dict else {}
        result = list(
            map(
                lambda x: data[x['name']] if x['name'] in data.keys() else None,
                columnsConf,
            )
        )
        return result
    rows = list(
        map(
            lambda x: getDatas(x),
            rows
        )    
    )
    return rows
def Export(
    filename,
    export_type = 'csv',
    datas = {
        'rows': [],
        'columnsConf': {},
        'title': None,
        'lang': None,
    },
    winMessage = {},
    loseMessage = {
        'type': responsesPossibles['unknown_error']['type'],
        'code': responsesPossibles['unknown_error']['code'],
        'status': responsesPossibles['unknown_error']['status'],
        'message' : 'Echec lors de l\'exportation du file',
    },
    returnException: bool = True,
) -> Response:
    returnException = deepcopy(returnException) if type(returnException) == bool else False
    filename = deepcopy(filename) if type(filename) in (list, tuple, dict, int, float, str, bool) else filename
    export_type = deepcopy(export_type) if type(export_type) in (list, tuple, dict, int, float, str, bool) else export_type
    datas = deepcopy(datas) if type(datas) in (list, tuple, dict, int, float, str, bool) else datas
    winMessage = deepcopy(winMessage) if type(winMessage) in (list, tuple, dict, int, float, str, bool) else winMessage
    loseMessage = deepcopy(loseMessage) if type(loseMessage) in (list, tuple, dict, int, float, str, bool) else loseMessage

    winMessage = winMessage if ( type(winMessage) == dict and 'type' in list(winMessage.keys()) and type(winMessage['type']) == str and len(winMessage['type']) > 0 and 'message' in list(winMessage.keys()) and type(winMessage['message']) == str and len(winMessage['message']) > 0 ) else {
        'type': responsesPossibles['good_action']['type'],
        'code': responsesPossibles['good_action']['code'],
        'status': responsesPossibles['good_action']['status'],
        'message' : 'Exportation des données au format \'{0}\' réalisée avec succès'
    }
    loseMessage = loseMessage if ( type(loseMessage) == dict and 'type' in list(loseMessage.keys()) and type(loseMessage['type']) == str and len(loseMessage['type']) > 0 and 'message' in list(loseMessage.keys()) and type(loseMessage['message']) == str and len(loseMessage['message']) > 0 ) else {
        'type': responsesPossibles['unknown_error']['type'],
        'code': responsesPossibles['unknown_error']['code'],
        'status': responsesPossibles['unknown_error']['status'],
        'message' : 'Echec lors de l\'exportation du file',
    }
    result = {}
    try:
        #config
        __filename = CleanName(filename)
        __filename = __filename if type(__filename) == str and len(__filename) > 0 else 'content'
        export_type = export_type if export_type in export_types else export_types[0]
        filename = Export__initFilename(filename, export_type = export_type)

        logo = EncodeFile(createStaticMediaPath('logo.app.png'))

        lang = getLang(datas['lang']) if (
            type(datas) == dict and
            'lang' in datas.keys() and
            type(datas['lang']) == str and
            len(datas['lang']) > 0
        ) else 'fr'
        
        langDatas = langdatas

        title = datas['title'] if (
            type(datas) == dict and
            'title' in datas.keys() and
            type(datas['title']) == str and
            len(datas['title']) > 0
        ) else 'liste des elements'

        columnsConf = datas['columnsConf'] if (
            type(datas) == dict and
            'columnsConf' in datas.keys() and
            type(datas['columnsConf']) == list
        ) else []
        columnsConf = Export__initColumnsConf(columnsConf)
        columns = Export__initColumns(columnsConf)
        
        # if DEBUG :
            # print("> file - export | export_type:: ", export_type)
            # print("> file - export | lang:: ", lang)
            # print("> file - export | title:: ", title)
            # print("> file - export | columnsConf:: ", columnsConf)
            # print("> file - export | columns:: ", columns)

        rows = datas['rows'] if (
            type(datas) == dict and
            'rows' in datas.keys() and
            (
                type(datas['rows']) == list or
                type(datas['rows']) == tuple
            )
        ) else {}
        jsonValue = deepcopy(rows)
        # if DEBUG :
            # print("> file - export | datas['rows'] (old):: ", datas['rows'])
        QRCode = Export_genQRCode(
            rows,
        )
        rows = Export__initRows(rows, columnsConf)
        
        # if DEBUG :
        #     print("> file - export | rows:: ", rows)

        resp = None
        if export_type == 'excel':
            # response = HttpResponse(
            #     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8',
            # )
            # response['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = __filename.capitalize()

            # Define some styles and formatting that will be later used for cells
            header_title_font = Font(name='Arial', bold=True, size=20)
            header_font = Font(name='Arial', bold=True)
            centered_alignment = Alignment(horizontal='center')
            border_bottom = Border(
                bottom=Side(border_style='medium', color='FF000000'),
            )
            wrapped_alignment = Alignment(
                vertical='top',
                wrap_text=True
            )
            fill = PatternFill(
                start_color="222222",
                end_color="eeeeee",
                fill_type='solid',
            )

            row_num = 0
            col_num = 1
            #init title
            """row_num = 1
            col_num = 1
            worksheet.merge_cells(
                start_row=row_num,
                start_column=1,
                end_row=row_num,
                end_column=(
                    len(columns) if type(columns) in [list, tuple] and len(columns) > 0 else 1
                )
            )
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = title
            cell.font = header_title_font
            cell.alignment = centered_alignment"""
            #init columns
            row_num = row_num + 1
            # row_num = row_num + 2
            col_num = 1
            for column in columns:
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column
                cell.font = header_font
                cell.border = border_bottom
                cell.alignment = centered_alignment
                #cell.fill = fill
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = 35
                col_num = col_num + 1
            #init rows
            row_num = row_num + 1
            for row in rows:
                col_num = 1
                for col_row in row:
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = col_row
                    #cell.style = cell_format
                    cell.alignment = wrapped_alignment
                    column_letter = get_column_letter(col_num)
                    column_dimensions = worksheet.column_dimensions[column_letter]
                    column_dimensions.width = 35
                    col_num = col_num + 1
                row_num = row_num + 1

            worksheet.freeze_panes = worksheet['A2']
            worksheet.sheet_properties.tabColor = 'ffffff'

            # workbook.save(filename)

            content = save_virtual_workbook(workbook)
            response = make_response(content)
            response.headers['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
            response.headers['AC-filename'] = filename
            response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8"
            resp = response
        elif export_type == 'pdf':
            html = render_template('pdf/list-data.template.html', 
                lang = lang,
                the = langDatas['le'][lang],
                datasLang = {},
                datenow = datetime.today().strftime("%d/%m/%Y, %H:%M:%S"),
                columns = columns,
                rows = rows,
                qrcode = QRCode,
                title = title,
                logo = logo,
                ese = eseDatas,
            )
            response = render_pdf(
                HTML(string=html),
                stylesheets=[
                    CSS(createStaticMediaPath('weasyprint','main.css')),
                    CSS(createStaticMediaPath('weasyprint','extend.css')),
                    CSS(createStaticMediaPath('weasyprint','list-data.css')),
                ],
                download_filename=filename,
            )
            resp = response
        elif export_type == 'csv':
            si = StringIO()
            writer = csv.writer(si)
            rows.insert(0, columns)
            for value in rows:
                writer.writerow(value)
            content = si.getvalue()
            response = make_response(content)
            response.headers['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
            response.headers['AC-filename'] = filename
            response.headers["Content-type"] = "text/csv; charset=utf-8"
            resp = response
        elif export_type == 'json':
            json_str = json.dumps(jsonValue)
            response = make_response(json_str)
            response.headers['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
            response.headers['AC-filename'] = filename
            response.headers["Content-type"] = "application/json; charset=utf-8"
            resp = response
        else:
            resp = jsonify(loseMessage)

        result = resp
    except Exception as err:
        if returnException == True:
            raise err
        else:
            code = str(type(err))
            msg = str(err)
            stack = str(traceback.format_exc())
            trace = sys.exc_info()[2]
            
            if DEBUG == True:
                log.error(stack)

            logging.getLogger("error_logger").error(traceback.format_exc())

            stackError = stack if DEBUG else None
            loseMessage['stack'] = stackError

            result = jsonify(loseMessage)
    return result

def Import(
    file: FileStorage,
    rows = {},
    columns = {},
    cleanData = (lambda x: x),
    schemas = {},
    returnException: bool = True,
):
    returnException = deepcopy(returnException) if type(returnException) == bool else False
    result = []
    invalidDatas = {}
    stackError = None
    try:
        #config
        columns = columns if type(columns) == dict else {
            'id': 0,
        }
        rows = rows if type(rows) == dict and 'begin' in rows.keys() and 'end' in rows.keys() else {
            "begin": 0,
            "end": 0
        }
        rows['begin'] = rows['begin'] if 'begin' in rows.keys() and type(rows['begin']) == int else None
        rows['end'] = rows['end'] if 'end' in rows.keys() and type(rows['end']) == int and rows['end'] != 0 else None

        schemas = schemas if type(schemas) == dict else {}
        
        cleanData = cleanData if (
            callable(cleanData)
        ) else (lambda x: x)
        
        # if DEBUG :
            # print("> file_manipulation - import - columns::", columns)
            # print("> file_manipulation - import - rows::", rows)
            # print("> file_manipulation - import - rows['begin']::", rows['begin'])
            # print("> file_manipulation - import - rows['end']::", rows['end'])
        
        def getDim2Tab_ExtractFileDatas(x):
            y = deepcopy(columns)
            y = y if type(y) == dict else {}
            result = {}
            # if DEBUG :
                # print("\t> getDim2Tab_ExtractFileDatas - y:: ", y)
            for index, (key, value) in enumerate(y.items()):
                # if DEBUG :
                    # print("\t\t> getDim2Tab_ExtractFileDatas - index:: ", index)
                    # print("\t\t> getDim2Tab_ExtractFileDatas - key:: ", key)
                    # print("\t\t> getDim2Tab_ExtractFileDatas - value:: ", value)
                    # print("\t\t> getDim2Tab_ExtractFileDatas - x:: ", x)
                result[key] = x[value] if type(value) == int and value < len(x) else None
            for prop in result:
                if type(result[prop]) == str : 
                    result[prop] = result[prop].lower()
            return result

        if file.filename.split('.')[-1] == 'csv' :
            result = pdDataframeToObject(pd.read_csv(file), isArray=True)
        elif file.filename.split('.')[-1] == 'json' :
            file.open()
            fileContent = file.read().decode('latin-1')
            result = json.loads(fileContent)
            result = list(
                map(
                    lambda x: list(x.values()),
                    result,
                )
            ) if type(result) in (list, tuple) else []
        elif file.filename.split('.')[-1] == 'xls' :
            result = pdDataframeToObject(pd.read_excel(file), isArray=True)
        elif file.filename.split('.')[-1] == 'xlsx' :
            result = pdDataframeToObject(pd.read_excel(file), isArray=True)
        if file.filename.split('.')[-1] in [ 'xls', 'xlsx', 'csv', 'json' ] :
            result = list(
                map(
                    lambda x: list(x.values()),
                    result,
                )
            )

        # if DEBUG :
            # print('> file_manipulation - import - brut result::', result)
        
        if rows['begin'] and rows['end']:
            result = result[ rows['begin']:rows['end'] ]
        elif rows['begin']:
            result = result[ rows['begin']: ]
        elif rows['end']:
            result = result[ :rows['end'] ]
        # if DEBUG :
            # print("> file_manipulation - import - result(step 1)::", result)
        result = list(
            map(
                lambda x: getDim2Tab_ExtractFileDatas(x),
                result
            )
        )
        result = list(
            map(
                lambda data: cleanData(data),
                result,
            ),
        )
        # if DEBUG :
            # print('> file_manipulation - import - result::', result)

        # validation
        invalidDatas = []
        for index, data in enumerate(result):
            dataID = {
                'index': index,
                'schemas': [],
            }
            dataIDIsInvalid = False
            for indexSchema, (keySchema, schema) in enumerate(schemas.items()):
                if(not(schema.isValid(data) == True)):
                    dataIDIsInvalid = True
                    dataID["schemas"].append(keySchema)
            if(dataIDIsInvalid == True):
                invalidDatas.append(dataID)
    except Exception as err:
        if returnException == True:
            raise err
        else:
            code = str(type(err))
            msg = str(err)
            stack = str(traceback.format_exc())
            trace = sys.exc_info()[2]
            
            if DEBUG == True:
                log.error(stack)

            logging.getLogger("error_logger").error(traceback.format_exc())
            """result = {
                'type': self.notifications['echec-extract-file']['type'],
                'message': self.notifications['echec-extract-file']['resp'][self.lang]
            }"""
            result = []
            invalidDatas = []
            stackError = stack if DEBUG else None

    return {
        'datas': result,
        'meta': {
            'invalid-datas': invalidDatas,
        },
        'notif': {
            'type': responsesPossibles['good_action']['type'],
            'code': responsesPossibles['good_action']['code'],
            'status': responsesPossibles['good_action']['status'],
            'message': responsesPossibles['good_action']['message']['fr'],
        } if len(invalidDatas) <= 0 else {
            'type': responsesPossibles['unknown_error']['type'],
            'code': responsesPossibles['unknown_error']['code'],
            'status': responsesPossibles['unknown_error']['status'],
            'message': responsesPossibles['unknown_error']['message']['fr'],
            'stack': stackError,
            # 'trace': sys.exc_info()[2],
        }
    }