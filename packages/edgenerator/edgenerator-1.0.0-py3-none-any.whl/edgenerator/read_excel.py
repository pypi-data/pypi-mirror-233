import openpyxl as xl
from edgenerator.state_classes import Pathway, State
from edgenerator.ed_class import EDGeneratorSettings

# o=======================================================o
#                        Functions
# o=======================================================o


def read_excelfile(excel_file_path: str, ed_settings: EDGeneratorSettings) -> EDGeneratorSettings:
    """
    Reads an excel file and returns a dictionary with the data

    Args:
        excel_file_path (str): path to the excel file
        ed_settings (EDGeneratorSettings): EDGeneratorSettings instance

    Returns:
        EDGeneratorSettings: EDGeneratorSettings instance

    Data structure of the Excel file:
    'BEGIN'
    'Settings'
    'energy_type' 'energy'
    'outputfile' 'test'
    ''
    'Paths' 'R' 'RC' 'TS' 'P' 'end'
    'Pathway1' -40 -30 -20 -10
    'Pathway2' -40 -30 -20 -10
    'Pathway3' -40 -30 -20 -10
    ...
    'END'

    Data structure of EDGeneratorSettings:
    'settings': ::EDProfile:: instance,
    'paths': [
        ::Pathway:: instance containing ::States:: instances
        ::Pathway:: instance containing ::States:: instances
        .
        .
        .
        ]

    """
    # Load whole workbook including different sheets
    wb_obj = xl.load_workbook(excel_file_path)

    # Get the active sheet (default is the first sheet)
    sheet_obj = wb_obj.active

    # General reading settings
    row_limit = 100
    column_limit = 100
    start_reading_index = -1

    # Find the 'BEGIN' to start reading
    for i_row in range(1, row_limit):
        if str(sheet_obj.cell(row=i_row, column=1).value).upper() == "BEGIN":
            start_reading_index = i_row
            break

    # SETTINGS READING BLOCK
    # This block reads the settings from the excel file
    # These entail type of energy, output file name, profile, etc.
    # The block ends with an "End" statement or when the "Paths" row is reached
    settings_block = False
    path_block = False
    headers = []

    i_row = start_reading_index
    j_col = 2
    while i_row < row_limit:
        # Loops over the rows and stops when it reaches the "paths" row
        col1 = sheet_obj.cell(row=i_row, column=1).value
        col2 = sheet_obj.cell(row=i_row, column=2).value
        if col1 == "end":
            break
        if str(col1).lower() == "settings":
            settings_block = True
        if str(col1).lower() == "paths":
            settings_block = False
            path_block = True
            while j_col < column_limit:
                # looping over the headers and stops at an "end" statement (column-wise)
                if str(sheet_obj.cell(row=i_row, column=j_col).value).lower() == "end":
                    break
                headers.append(sheet_obj.cell(row=i_row, column=j_col).value)
                j_col += 1
            i_row += 1
            break
        if settings_block and col1 is not None:
            if col1.lower() == "energy_type" or col1.lower() == "energy" or col1.lower() == "energy type":
                ed_settings.settings.energy_type = col2
            elif col1.lower() == "outputfile" or col1.lower() == "output file" or col1.lower() == "output_file" or col1.lower() == "output":
                ed_settings.settings.outputfile = col2
            else:
                pass
        i_row += 1

    # PATHS READING BLOCK
    # This block reads the pathways from the excel file and stores them in a list
    # These entail the name of pathways and the states (instances of ::State::)
    # The block ends with an "End" statement (row-wise)
    if not path_block:
        raise ValueError("Error: no 'Paths' found in the excel file")

    # Loop over the rows and columns
    while i_row < row_limit:
        if str(sheet_obj.cell(row=i_row, column=1).value) == "END":
            break
        pathway_name = sheet_obj.cell(row=i_row, column=1).value
        pathway = Pathway(name=pathway_name)
        for j_col in range(2, len(headers)+2):
            cell_value = sheet_obj.cell(row=i_row, column=j_col).value
            if cell_value is not None:
                state = State()
                state.label = f"{i_row+2-len(headers)}_{headers[j_col-2]}"
                state.column = j_col-1
                state.energy = cell_value
                pathway.add_state(state)
            j_col += 1
        ed_settings.paths.append(pathway)
        i_row += 1

    return ed_settings
