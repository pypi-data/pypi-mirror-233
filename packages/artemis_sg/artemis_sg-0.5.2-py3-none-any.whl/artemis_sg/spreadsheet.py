import logging
import math
import os
import re
from copy import copy
from inspect import getsourcefile

from googleapiclient.discovery import build
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from PIL import Image as PIL_Image
from PIL import UnidentifiedImageError

import artemis_sg.app_creds as app_creds
import artemis_sg.vendor as vendor
from artemis_sg.config import CFG

MODULE = os.path.splitext(os.path.basename(__file__))[0]


def get_worksheet(wb_obj, worksheet):
    if not worksheet:
        ws = wb_obj.worksheets[0]
    else:
        ws = wb_obj[worksheet]
    return ws


def sheet_image(vendor_code, workbook, worksheet, image_directory, out):
    namespace = f"{MODULE}.{sheet_image.__name__}"

    image_row_height = CFG["asg"]["spreadsheet"]["sheet_image"]["image_row_height"]
    image_col_width = CFG["asg"]["spreadsheet"]["sheet_image"]["image_col_width"]
    isbn_col_width = CFG["asg"]["spreadsheet"]["sheet_image"]["isbn_col_width"]
    max_col_width = CFG["asg"]["spreadsheet"]["sheet_image"]["max_col_width"]
    col_buffer = CFG["asg"]["spreadsheet"]["sheet_image"]["col_buffer"]

    # get vendor info from database
    logging.debug(f"{namespace}: Instantiate vendor.")
    vendr = vendor.Vendor(vendor_code)
    vendr.set_vendor_data()

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")

    # Insert column "A" for "ISBN"
    # Insert column "B" for "Image"
    # Insert column "C" for "Order"
    ws.insert_cols(1)
    ws.insert_cols(1)
    ws.insert_cols(1)
    ws["B1"] = "Image"
    ws["C1"] = "Order"

    # Move ISBN colum to "A"
    # Find ISBN column
    row01 = ws[1]
    for cell in row01:
        if type(cell.value) is str and cell.value.upper() == isbn_key.upper():
            # Copy formatting
            h_font = copy(cell.font)
            h_border = copy(cell.border)
            h_fill = copy(cell.fill)
            h_number_format = copy(cell.number_format)
            h_protection = copy(cell.protection)
            h_alignment = copy(cell.alignment)
            break
    # Move it to A
    isbn_idx = cell.column
    ws.move_range(
        f"{cell.column_letter}{cell.row}:{cell.column_letter}{ws.max_row}",
        cols=-(cell.column - 1),
    )
    ws.delete_cols(isbn_idx)

    # Copy ISBN header format to B1, C1
    cell = ws["A1"]
    if cell.has_style:
        # fmt: off
        ws["B1"].font          = ws["C1"].font          = h_font  # noqa: E221, E501
        ws["B1"].border        = ws["C1"].border        = h_border  # noqa: E221, E501
        ws["B1"].fill          = ws["C1"].fill          = h_fill  # noqa: E221, E501
        ws["B1"].number_format = ws["C1"].number_format = h_number_format  # noqa: E501
        ws["B1"].protection    = ws["C1"].protection    = h_protection  # noqa: E221, E501
        ws["B1"].alignment     = ws["C1"].alignment     = h_alignment  # noqa: E221, E501
        # fmt: on

    # Create column widths
    dim_holder = DimensionHolder(worksheet=ws)

    # Set column A isbn_col_width for ISBN numbers
    dim_holder["A"] = ColumnDimension(
        ws, index="A", width=math.ceil(isbn_col_width * col_buffer)
    )
    # Set column B image_col_width for images
    dim_holder["B"] = ColumnDimension(ws, index="B", width=image_col_width)

    # Dynamically set remaining columns
    for col in range(3, ws.max_column + 1):
        col_letter = get_column_letter(col)
        width = max(len(str(cell.value)) for cell in ws[col_letter]) * col_buffer
        if width > max_col_width:
            width = max_col_width
        dim_holder[col_letter] = ColumnDimension(ws, index=col_letter, width=width)

    ws.column_dimensions = dim_holder

    # Prepare column "B" for "Image"
    colB = ws["B"]
    for cell in colB:
        # Format to center content
        cell.alignment = Alignment(horizontal="center")

    # Insert images in column 2, (i.e. "B")
    for row in ws.iter_rows(min_row=2, max_col=2):
        isbn_cell, image_cell = row
        if isbn_cell.value:
            isbn = isbn_cell.value
            if isinstance(isbn, float):
                isbn = int(isbn)
            elif isinstance(isbn, str):
                m = re.search('="(.*)"', isbn)
                if m:
                    isbn = m.group(1)
            try:
                isbn = str(isbn).strip()
            except Exception as e:
                logging.error(f"{namespace}: Err reading isbn '{isbn}', err: '{e}'")
                isbn = ""
        else:
            isbn = ""
        # Set row height
        row_dim = ws.row_dimensions[image_cell.row]
        row_dim.height = image_row_height

        # Insert image into cell
        filename = f"{isbn}.jpg"
        filepath = os.path.join(image_directory, filename)
        logging.debug(f"{namespace}: Attempting to insert '{filepath}'.")
        if os.path.isfile(filepath):
            img = Image(filepath)
            ws.add_image(img, f"{image_cell.column_letter}{image_cell.row}")
            logging.info(f"{namespace}: Inserted '{filepath}'.")

    # Save workbook
    wb.save(out)


def get_order_items(vendor_code, workbook, worksheet):
    namespace = f"{MODULE}.{get_order_items.__name__}"

    order_items = []
    # get vendor info from database
    logging.debug(f"{namespace}: Instantiate vendor.")
    vendr = vendor.Vendor(vendor_code)
    vendr.set_vendor_data()

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")

    # Find Isbn and Order column letters
    row01 = ws[1]
    for cell in row01:
        if cell.value == isbn_key:
            isbn_column_letter = cell.column_letter
        if cell.value == "Order":
            order_column_letter = cell.column_letter

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter == isbn_column_letter:
                isbn_cell = cell
            if cell.column_letter == order_column_letter:
                order_cell = cell
        # Validate ISBN
        if not isbn_cell.value:
            continue
        else:
            isbn = isbn_cell.value
            if isinstance(isbn, str):
                m = re.search('="(.*)"', isbn)
                if m:
                    isbn = m.group(1)
            try:
                isbn = str(int(isbn)).strip()
            except Exception as e:
                logging.error(f"{namespace}: Err reading isbn '{isbn}', err: '{e}'")
                continue
        # Validate Order Qty
        if not order_cell.value:
            continue
        else:
            qty = order_cell.value
            try:
                qty = str(int(qty)).strip()
            except Exception as e:
                logging.error(f"{namespace}: Err reading Order qty '{qty}', err: '{e}'")
                continue
        order_items.append((isbn, qty))

    return order_items


def mkthumbs(image_directory):
    namespace = f"{MODULE}.{mkthumbs.__name__}"

    thumb_width = CFG["asg"]["spreadsheet"]["mkthumbs"]["width"]
    thumb_height = CFG["asg"]["spreadsheet"]["mkthumbs"]["height"]

    here = os.path.dirname(getsourcefile(lambda: 0))
    data = os.path.abspath(os.path.join(here, "data"))
    logo = os.path.join(data, "artemis_logo.png")
    logging.debug(f"{namespace}: Found image for thumbnail background at '{logo}'")
    sub_dir = "thumbnails"
    back = PIL_Image.open(logo)
    thumb_dir = os.path.join(image_directory, sub_dir)
    logging.debug(f"{namespace}: Defining thumbnail directory as '{thumb_dir}'")
    if not os.path.isdir(thumb_dir):
        logging.debug(f"{namespace}: Creating directory '{thumb_dir}'")
        os.mkdir(thumb_dir)
        if os.path.isdir(thumb_dir):
            logging.info(f"{namespace}: Successfully created directory '{thumb_dir}'")
        else:
            logging.error(
                f"{namespace}: Failed to create directory '{thumb_dir}'. Aborting."
            )
            raise Exception
    files = os.listdir(image_directory)
    for f in files:
        # Valid files are JPG or PNG that are not supplemental images.
        image = re.match(r"^.+\.(?:jpg|png)$", f)
        if not image:
            continue
        # Supplemental images have a "-[0-9]+" suffix before the file type.
        # AND a file without that suffix exists int he image_directory.
        suffix = re.match(r"(^.+)-[0-9]+(\.(?:jpg|png))$", f)
        if suffix:
            primary = suffix.group(1) + suffix.group(2)
            primary_path = os.path.join(image_directory, primary)
            if os.path.isfile(primary_path):
                continue
        thumb_file = os.path.join(thumb_dir, f)
        # don't remake thumbnails
        if os.path.isfile(thumb_file):
            continue
        bk = back.copy()
        try:
            file_path = os.path.join(image_directory, f)
            fg = PIL_Image.open(file_path)
        except UnidentifiedImageError:
            logging.error(f"{namespace}: Err reading '{f}', deleting '{file_path}'")
            os.remove(file_path)
            continue
        fg.thumbnail((thumb_width, thumb_height))
        size = (int((bk.size[0] - fg.size[0]) / 2), int((bk.size[1] - fg.size[1]) / 2))
        bk.paste(fg, size)
        logging.debug(f"{namespace}: Attempting to save thumbnail '{thumb_file}'")
        bkn = bk.convert("RGB")
        bkn.save(thumb_file)
        logging.info(f"{namespace}: Successfully created thumbnail '{thumb_file}'")


def get_sheet_data(workbook, worksheet=None):
    namespace = f"{MODULE}.{get_sheet_data.__name__}"
    #########################################################################
    # Try to open sheet_id as an Excel file
    sheet_data = []
    try:
        wb = load_workbook(workbook)
        ws = get_worksheet(wb, worksheet)
        for row in ws.values:
            sheet_data.append(row)
    except (FileNotFoundError, InvalidFileException):
        #########################################################################
        # Google specific stuff
        # authenticate to google sheets
        logging.info(f"{namespace}: Authenticating to google api.")
        creds = app_creds.app_creds()
        SHEETS = build("sheets", "v4", credentials=creds)
        # get sheet data
        if not worksheet:
            sheets = (
                SHEETS.spreadsheets()
                .get(spreadsheetId=workbook)
                .execute()
                .get("sheets", "")
            )
            ws = sheets.pop(0).get("properties", {}).get("title")
        else:
            ws = worksheet
        sheet_data = (
            SHEETS.spreadsheets()
            .values()
            .get(range=ws, spreadsheetId=workbook)
            .execute()
            .get("values")
        )
        #########################################################################
    return sheet_data
